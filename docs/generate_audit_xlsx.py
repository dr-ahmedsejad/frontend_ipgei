"""
Génère le questionnaire d'audit SI au format Excel (.xlsx).

Sortie : docs/audit_questionnaire_si.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# Contenu
# ---------------------------------------------------------------------------

AXES = [
    {
        "num": 1,
        "titre": "Gouvernance & stratégie SI",
        "siga_global": "Hors périmètre",
        "questions": [
            ("1.1", "Schéma directeur SI formalisé ? Date de dernière révision ?", "Hors périmètre", ""),
            ("1.2", "Responsable SI désigné (DSI interne / prestataire / volontaire) ?", "Hors périmètre", ""),
            ("1.3", "Budget annuel SI et part dédiée aux applicatifs métier ?", "Hors périmètre", ""),
            ("1.4", "Cartographie applicative à jour ?", "Hors périmètre", ""),
            ("1.5", "Comité de pilotage SI (fréquence, composition) ?", "Hors périmètre", ""),
        ],
    },
    {
        "num": 2,
        "titre": "Infrastructure, sécurité & conformité",
        "siga_global": "Partiel",
        "questions": [
            ("2.1", "Hébergement (on-premise / cloud / mixte) et prestataire ?", "Hors périmètre", ""),
            ("2.2", "Politique de sauvegarde (fréquence, externalisation, tests de restauration) ?", "Hors périmètre", ""),
            ("2.3", "Authentification : SSO, AD/LDAP, comptes locaux ?", "Partiel", "comptes locaux + JWT refresh + logout inactivité 20 min"),
            ("2.4", "RBAC formalisé (matrice rôles × modules) ?", "Natif", "auth-roles.ts — admin / IT / DE / scolarite / dept / prof / etudiant"),
            ("2.5", "Mécanisme de déblocage de compte tracé ?", "Natif", "dashboard/deblocage"),
            ("2.6", "Conformité RGPD / loi locale données personnelles ?", "Hors périmètre", ""),
            ("2.7", "Politique de mot de passe (rotation, complexité, MFA) ?", "Partiel", "changement mot de passe oui, MFA non"),
        ],
    },
    {
        "num": 3,
        "titre": "Audit & traçabilité",
        "siga_global": "Natif",
        "questions": [
            ("3.1", "Journal d'audit centralisé append-only ?", "Natif", "apps/audit/ — AuditLog append-only garanti côté modèle"),
            ("3.2", "Timeline par entité (qui a touché quoi, quand) ?", "Natif", "/api/v1/audit/by-entity/?model=X&object_id=Y"),
            ("3.3", "Stats globales d'activité par action / modèle ?", "Natif", "/api/v1/audit/stats/"),
            ("3.4", "Export CSV des logs ?", "Natif", "/api/v1/audit/export/ (streamé)"),
            ("3.5", "Restriction d'accès : un prof voit-il l'audit d'un autre ?", "Natif", "Admin/IT global, autres = by-entity uniquement"),
            ("3.6", "Historique des modifications visible côté utilisateur ?", "Natif", "dashboard/historique"),
        ],
    },
    {
        "num": 4,
        "titre": "Référentiels académiques",
        "siga_global": "Natif",
        "questions": [
            ("4.1", "Référentiel filières, départements, niveaux, modules ?", "Natif", "parametres/, scolarite/filieres, scolarite/departements"),
            ("4.2", "Versionnage par année universitaire ?", "Natif", "parametres/annees, scoping annee_universitaire partout"),
            ("4.3", "Gestion semestres, semaines, créneaux ?", "Natif", "parametres/semestres, semaines, creneaux"),
            ("4.4", "Référentiel jours de la semaine paramétrable ?", "Natif", "parametres/jours"),
            ("4.5", "Référentiel salles avec capacité ?", "Natif", "parametres/salles, salles/"),
            ("4.6", "Calendrier adapté (Ramadan, événements locaux) ?", "Natif", "parametres/ramadan"),
            ("4.7", "Périodes de réclamation paramétrables ?", "Natif", "parametres/periodes-reclamation"),
        ],
    },
    {
        "num": 5,
        "titre": "Multi-établissement",
        "siga_global": "Natif",
        "questions": [
            ("5.1", "Plusieurs institutions partagent-elles un même SI ?", "Natif", "parametres/institutions, institution/"),
            ("5.2", "Isolation stricte des données par institution (FK partout) ?", "Natif", "scoping institution sur emplois/suivi/vacation/documents"),
            ("5.3", "Consolidation cross-institution pour direction du groupe ?", "Partiel", "stats par institution, agrégation groupe à confirmer"),
            ("5.4", "Identifiant unique étudiant / enseignant au niveau groupe ?", "Partiel", "identifiants internes par institution"),
        ],
    },
    {
        "num": 6,
        "titre": "Pré-inscriptions & admissions",
        "siga_global": "Natif",
        "questions": [
            ("6.1", "Pré-inscription en ligne (lien public) ?", "Natif", "inscriptions/preinscriptions/[token]"),
            ("6.2", "Token sécurisé / lien unique par candidat ?", "Natif", "route [token]"),
            ("6.3", "Workflow validation candidature → inscription administrative ?", "Natif", "inscriptions/administratives"),
            ("6.4", "Dérogations d'inscription tracées ?", "Natif", "inscriptions/derogations"),
            ("6.5", "Import en masse étudiants (CSV/Excel) ?", "Natif", "scolarite/etudiants/importer"),
        ],
    },
    {
        "num": 7,
        "titre": "Vie étudiante & scolarité",
        "siga_global": "Natif",
        "questions": [
            ("7.1", "Dossier étudiant centralisé (état civil, parcours, documents) ?", "Natif", "scolarite/etudiants/[id], documents/etudiant/[id]"),
            ("7.2", "Recherche multi-critères étudiants ?", "Natif", "scolarite/etudiants/chercher"),
            ("7.3", "Inscription pédagogique (choix modules) séparée de l'administrative ?", "Natif", "inscriptions/pedagogiques"),
            ("7.4", "Suivi de la progression académique (semestre/année) ?", "Natif", "scolarite/progressions, dashboard/avancement"),
            ("7.5", "Gestion des comptes étudiants (création, blocage) ?", "Natif", "scolarite/etudiants/comptes"),
            ("7.6", "Portail étudiant (notes, EDT, absences, documents) ?", "Natif", "portail/ (emploi, notes, releve, absences, documents, reclamations)"),
        ],
    },
    {
        "num": 8,
        "titre": "Emplois du temps",
        "siga_global": "Natif",
        "questions": [
            ("8.1", "Conception EDT centralisée ou décentralisée par département ?", "Natif", "emplois/, em/"),
            ("8.2", "Détection automatique des conflits (prof / salle / classe) ?", "Natif", "scoping FK Phase 5 garantit cohérence référentielle"),
            ("8.3", "Gestion des séances de rattrapage ?", "Natif", "via suivi/ (type_seance)"),
            ("8.4", "Archivage EDT par semaine / année ?", "Natif", "EmploisArchive"),
            ("8.5", "Diffusion (portail prof + portail étudiant) ?", "Natif", "portail/emploi, enseignant/emploi"),
        ],
    },
    {
        "num": 9,
        "titre": "Ressources enseignants & vacations",
        "siga_global": "Natif",
        "questions": [
            ("9.1", "Annuaire enseignants distinguant permanents / vacataires ?", "Natif", "prof/, enseignant/"),
            ("9.2", "Suivi des charges horaires par enseignant ?", "Natif", "suivi/charges"),
            ("9.3", "Détail enseignements par prof ?", "Natif", "enseignant/detail-enseignements"),
            ("9.4", "Génération états de paiement vacations ?", "Natif", "vacation/, enseignant/vacations, dashboard/payement"),
            ("9.5", "Export bancaire (virements) ?", "Natif", "dashboard/banque"),
            ("9.6", "Intégration comptabilité générale ?", "Partiel", "export seulement, pas d'API compta"),
        ],
    },
    {
        "num": 10,
        "titre": "Suivi pédagogique & présences",
        "siga_global": "Natif",
        "questions": [
            ("10.1", "Pointage de séance (enseignant présent, séance effective) ?", "Natif", "SuiviePointage, M2M pointage_departements"),
            ("10.2", "Saisie absences étudiants par séance / par salle ?", "Natif", "absences/saisir/salle/[suiviId]"),
            ("10.3", "Import en masse des absences (rétro-saisie) ?", "Natif", "absences/importer"),
            ("10.4", "Gestion des justificatifs avec workflow ?", "Natif", "absences/justificatifs"),
            ("10.5", "Saisie via portail enseignant ?", "Natif", "enseignant/suivi"),
            ("10.6", "Alertes seuil d'absences automatiques ?", "Partiel", "à confirmer dans notifications/"),
        ],
    },
    {
        "num": 11,
        "titre": "Évaluations, délibérations & qualité examens",
        "siga_global": "Natif",
        "questions": [
            ("11.1", "Saisie notes par enseignant via portail ?", "Natif", "evaluations/notes"),
            ("11.2", "Saisie anonyme (anonymat des copies) ?", "Natif", "evaluations/notes/saisie-anonymat"),
            ("11.3", "Sessions normale / rattrapage distinguées ?", "Natif", "SessionEvaluation (SN/SR)"),
            ("11.4", "Pondération CC/TP/Examen paramétrable ?", "Natif", "evaluations/ponderation, types CC/TP/EXAM"),
            ("11.5", "Compensation modulaire et semestrielle automatisée ?", "Natif", "codes statut V / VCI / VCS / R / NV / NVO / E"),
            ("11.6", "Délibérations avec PV générés ?", "Natif", "evaluations/deliberations/[id]/pv"),
            ("11.7", "Rachats jury tracés ?", "Natif", "evaluations/rachats, statut R (Rachat jury)"),
            ("11.8", "Décisions annuelles codifiées (passage / redoublement / exclusion / année blanche) ?", "Natif", "5 décisions normalisées"),
            ("11.9", "Articles réglementaires référencés dans les codes (Art. 12-23) ?", "Natif", "légende validée docs/deliberation.md"),
        ],
    },
    {
        "num": 12,
        "titre": "Stages & insertion professionnelle",
        "siga_global": "Natif",
        "questions": [
            ("12.1", "Conventions de stage numérisées ?", "Natif", "stages/conventions, ConventionStage"),
            ("12.2", "Statuts workflow (brouillon → soumise → en cours → terminée) ?", "Natif", "5 statuts standard"),
            ("12.3", "Distinction stage / PFE ?", "Natif", "flag est_pfe sur convention"),
            ("12.4", "Tuteur académique + tuteur entreprise tracés ?", "Natif", "FK Prof + champ tuteur_entreprise"),
            ("12.5", "Évaluation tri-axiale (entreprise / rapport / soutenance) ?", "Natif", "EvaluationStage"),
            ("12.6", "Jury soutenance M2M (plusieurs membres) ?", "Natif", "jury ManyToManyField"),
            ("12.7", "Validation PFE séparée de la note finale ?", "Natif", "est_valide_pfe"),
            ("12.8", "Classement / palmarès stages ?", "Natif", "stages/classement"),
            ("12.9", "Dérogations médicales avec justificatif fichier ?", "Natif", "stages/derogations, DerogationMedicale"),
        ],
    },
    {
        "num": 13,
        "titre": "Réclamations encadrées",
        "siga_global": "Natif",
        "questions": [
            ("13.1", "Réclamations étudiants tracées (absence / note / autre) ?", "Natif", "3 types : absence, note, autre"),
            ("13.2", "Fenêtres temporelles d'ouverture des réclamations ?", "Natif", "PeriodeReclamation paramétrable (SN/SR, Impairs/Pairs)"),
            ("13.3", "Réclamation rattachée à l'objet contesté (séance / note / session) ?", "Natif", "FK optionnelles presence / inscription_element / session_evaluation"),
            ("13.4", "Pièce jointe justificative ?", "Natif", "justificatif FileField"),
            ("13.5", "Workflow traitement (soumise → en_cours → acceptée/rejetée) ?", "Natif", "4 statuts + traceur traitee_par + dates"),
            ("13.6", "Visibilité côté étudiant (portail) ?", "Natif", "portail/reclamations"),
            ("13.7", "Visibilité côté enseignant ?", "Natif", "enseignant/reclamations"),
        ],
    },
    {
        "num": 14,
        "titre": "Documents, communication & reporting",
        "siga_global": "Natif",
        "questions": [
            ("14.1", "Génération automatique de documents (attestations, relevés) ?", "Natif", "scolarite/documents, documents/registre"),
            ("14.2", "Dossier documentaire par étudiant ?", "Natif", "documents/etudiant/[id]"),
            ("14.3", "Téléchargement via portail étudiant ?", "Natif", "portail/documents, portail/releve"),
            ("14.4", "Notifications internes (intra-application) ?", "Natif", "notifications/, dashboard/notifications"),
            ("14.5", "Notifications email / SMS sortantes ?", "Hors périmètre", "canaux externes à confirmer"),
            ("14.6", "Tableaux de bord direction (statistiques) ?", "Natif", "statistiques/semestres, statistiques/profs"),
            ("14.7", "Reporting réglementaire vers tutelle (export) ?", "Partiel", "exports CSV ponctuels, format ministériel à standardiser"),
            ("14.8", "Signature électronique des documents officiels ?", "Hors périmètre", ""),
            ("14.9", "GED avec archivage légal (durée, format) ?", "Partiel", "stockage fichiers oui, politique d'archivage non formalisée"),
        ],
    },
    {
        "num": 15,
        "titre": "Site web institutionnel & présence numérique",
        "siga_global": "Hors périmètre",
        "questions": [
            ("15.1", "Site web institutionnel à jour ? Date de la dernière refonte ?", "Hors périmètre", "CMS externe"),
            ("15.2", "CMS utilisé (WordPress, Drupal, Strapi, Next.js custom…) maintenu et à jour ?", "Hors périmètre", "externe"),
            ("15.3", "Site responsive mobile (consulté principalement sur smartphone par les candidats) ?", "Hors périmètre", "externe"),
            ("15.4", "Site multilingue (FR / AR / EN selon contexte) ?", "Hors périmètre", "externe"),
            ("15.5", "Accessibilité numérique (RGAA / WCAG 2.1 niveau AA) ?", "Hors périmètre", "externe"),
            ("15.6", "Catalogue des formations à jour avec maquettes / débouchés / fiches filières ?", "Partiel", "données SIGA exportables (filieres, modules)"),
            ("15.7", "Calendrier universitaire publié (vacances, examens, rentrées) ?", "Partiel", "calendrier interne SIGA réutilisable"),
            ("15.8", "Actualités & agenda événementiel mis à jour régulièrement (≥ mensuel) ?", "Hors périmètre", "externe"),
            ("15.9", "Formulaire de pré-candidature en ligne intégré au SI (pas de re-saisie) ?", "Natif", "inscriptions/preinscriptions/[token]"),
            ("15.10", "Annuaire enseignants / responsables / services consultable depuis le site ?", "Partiel", "données SIGA réutilisables"),
            ("15.11", "Lien vers portails (étudiant, enseignant) depuis page d'accueil ?", "Partiel", "lien externe vers SIGA"),
            ("15.12", "Single Sign-On (SSO) depuis le site → portail SI (une seule authentification) ?", "Hors périmètre", "auth SIGA séparée aujourd'hui"),
            ("15.13", "Brochures téléchargeables (PDF par filière) à jour ?", "Hors périmètre", "externe"),
            ("15.14", "Témoignages diplômés / alumni mis en avant ?", "Hors périmètre", "externe"),
            ("15.15", "Partenariats académiques & entreprises affichés ?", "Hors périmètre", "externe"),
            ("15.16", "Espace presse / publications scientifiques / rapports d'activité ?", "Hors périmètre", "externe"),
            ("15.17", "Présence et activité sur réseaux sociaux (Facebook, LinkedIn, YouTube, Instagram, TikTok) ?", "Hors périmètre", "externe"),
            ("15.18", "Newsletter / mailing institutionnel actif (fréquence, taux d'ouverture suivi) ?", "Hors périmètre", "externe"),
            ("15.19", "Mentions légales + politique cookies + politique vie privée (RGPD / loi locale) ?", "Hors périmètre", "externe"),
            ("15.20", "Statistiques de fréquentation suivies (Google Analytics, Matomo) ?", "Hors périmètre", "externe"),
            ("15.21", "Référencement SEO suivi (positionnement requêtes clés) ?", "Hors périmètre", "externe"),
            ("15.22", "Hébergement (prestataire, SLA, sauvegarde, certificat SSL) documenté ?", "Hors périmètre", "externe"),
            ("15.23", "Charte graphique cohérente entre site, portail SIGA et documents officiels ?", "Partiel", "logos & couleurs partageables"),
        ],
    },
]


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="006633")  # vert SIGA

SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="555555")

AXIS_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
AXIS_FILL = PatternFill("solid", fgColor="006633")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4A7C2E")

ROW_ALT_FILL = PatternFill("solid", fgColor="F2F8F2")

SIGA_NATIF_FILL = PatternFill("solid", fgColor="C6EFCE")  # vert clair
SIGA_PARTIEL_FILL = PatternFill("solid", fgColor="FFEB9C")  # jaune
SIGA_HORS_FILL = PatternFill("solid", fgColor="FFC7CE")  # rouge clair

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def siga_fill(value: str) -> PatternFill | None:
    if value == "Natif":
        return SIGA_NATIF_FILL
    if value == "Partiel":
        return SIGA_PARTIEL_FILL
    if value == "Hors périmètre":
        return SIGA_HORS_FILL
    return None


# ---------------------------------------------------------------------------
# Construction du classeur
# ---------------------------------------------------------------------------

def build_questionnaire_sheet(ws):
    ws.title = "Questionnaire"

    # En-tête principal
    ws.merge_cells("A1:F1")
    ws["A1"] = "Questionnaire d'audit SI — Établissements supérieurs"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = "Auteur : Dr. Ahmed SEJAD · Date : 2026-05-21 · Version : v2"
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Légende
    ws.merge_cells("A3:F3")
    ws["A3"] = (
        "Notation : 0=Inexistant · 1=Manuel · 2=Tableurs · 3=Logiciel partiel · "
        "4=Intégré · 5=Intégré + portail   |   Couverture SIGA : Natif / Partiel / Hors périmètre"
    )
    ws["A3"].font = Font(italic=True, size=10, color="555555")
    ws["A3"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # En-têtes de colonnes
    headers = ["#", "Question", "Note (0-5)", "Observations", "Couvert par SIGA", "Module / Référence SIGA"]
    header_row = 5
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 28

    # Largeurs de colonnes
    widths = {"A": 7, "B": 60, "C": 11, "D": 35, "E": 18, "F": 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Validation note 0-5
    dv_note = DataValidation(
        type="whole",
        operator="between",
        formula1=0,
        formula2=5,
        showErrorMessage=True,
        errorTitle="Note invalide",
        error="La note doit être un entier entre 0 et 5.",
    )
    ws.add_data_validation(dv_note)

    row = header_row + 1
    for axe in AXES:
        # Ligne de titre de l'axe
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=f"Axe {axe['num']} — {axe['titre']}  ·  SIGA : {axe['siga_global']}")
        cell.font = AXIS_FONT
        cell.fill = AXIS_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 24
        row += 1

        for i, (num, question, siga, ref) in enumerate(axe["questions"]):
            alt = (i % 2 == 1)
            values = [num, question, None, "", siga, ref]
            for col_idx, val in enumerate(values, start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.border = BORDER
                c.alignment = WRAP_TOP if col_idx in (2, 4, 6) else WRAP_CENTER
                if alt and col_idx != 5:
                    c.fill = ROW_ALT_FILL

            # Coloration cellule SIGA
            fill = siga_fill(siga)
            if fill is not None:
                ws.cell(row=row, column=5).fill = fill

            # Data validation sur la cellule Note
            dv_note.add(ws.cell(row=row, column=3))
            row += 1

    # Freeze panes sous l'en-tête
    ws.freeze_panes = "A6"


def build_synthese_sheet(ws):
    ws.title = "Synthèse par axe"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Synthèse par axe — après dépouillement"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Axe", "Couverture (0-5)", "Maturité (0-5)", "Risque (faible/moyen/élevé)", "Opportunité SIGA (Oui/Partiel/Non)"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[3].height = 30

    widths = {"A": 50, "B": 18, "C": 18, "D": 28, "E": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    dv_note = DataValidation(
        type="whole", operator="between", formula1=0, formula2=5,
        showErrorMessage=True, errorTitle="Note invalide",
        error="La note doit être un entier entre 0 et 5.",
    )
    ws.add_data_validation(dv_note)

    dv_risque = DataValidation(
        type="list", formula1='"faible,moyen,élevé"',
        showErrorMessage=True, errorTitle="Valeur invalide",
        error="Choisir parmi : faible / moyen / élevé",
    )
    ws.add_data_validation(dv_risque)

    dv_opportunite = DataValidation(
        type="list", formula1='"Oui,Partiel,Non"',
        showErrorMessage=True, errorTitle="Valeur invalide",
        error="Choisir parmi : Oui / Partiel / Non",
    )
    ws.add_data_validation(dv_opportunite)

    row = 4
    for i, axe in enumerate(AXES):
        alt = (i % 2 == 1)
        label = f"Axe {axe['num']} — {axe['titre']}"
        ws.cell(row=row, column=1, value=label).alignment = WRAP_TOP
        for col_idx in range(1, 6):
            c = ws.cell(row=row, column=col_idx)
            c.border = BORDER
            if alt:
                c.fill = ROW_ALT_FILL
        dv_note.add(ws.cell(row=row, column=2))
        dv_note.add(ws.cell(row=row, column=3))
        dv_risque.add(ws.cell(row=row, column=4))
        dv_opportunite.add(ws.cell(row=row, column=5))
        row += 1

    ws.freeze_panes = "A4"


def build_etablissement_sheet(ws):
    ws.title = "Établissement"

    ws.merge_cells("A1:B1")
    ws["A1"] = "Identification de l'établissement audité"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    fields = [
        "Nom de l'établissement",
        "Type (public / privé / mixte)",
        "Effectif étudiants",
        "Effectif enseignants (permanents + vacataires)",
        "Nombre de filières",
        "Nombre de départements",
        "Nombre de sites / campus",
        "Tutelle",
        "Nom du répondant principal",
        "Fonction",
        "Date de l'audit",
        "Auditeur",
    ]
    defaults = {
        "Auditeur": "Dr. Ahmed SEJAD",
    }

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 50

    for i, label in enumerate(fields, start=3):
        c_label = ws.cell(row=i, column=1, value=label)
        c_label.font = Font(bold=True)
        c_label.fill = PatternFill("solid", fgColor="F2F8F2")
        c_label.border = BORDER
        c_label.alignment = WRAP_TOP

        c_val = ws.cell(row=i, column=2, value=defaults.get(label, ""))
        c_val.border = BORDER
        c_val.alignment = WRAP_TOP


def build_legende_sheet(ws):
    ws.title = "Légende"

    ws.merge_cells("A1:B1")
    ws["A1"] = "Légende & mode d'emploi"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    sections = [
        ("Grille de notation", None),
        ("0", "Inexistant — aucun processus"),
        ("1", "Manuel / papier uniquement"),
        ("2", "Tableurs (Excel) en local, pas de partage structuré"),
        ("3", "Logiciel partiel — un module isolé, pas d'intégration"),
        ("4", "Logiciel intégré — référentiels partagés entre modules"),
        ("5", "Intégré + portail utilisateur (étudiant / enseignant)"),
        ("", ""),
        ("Couverture SIGA", None),
        ("Natif", "SIGA couvre la question avec un module dédié"),
        ("Partiel", "SIGA couvre une partie, complément externe nécessaire"),
        ("Hors périmètre", "SIGA ne traite pas ce besoin"),
        ("", ""),
        ("Synthèse SIGA globale", None),
        ("85 %", "du périmètre métier d'un établissement supérieur couvert nativement par SIGA (axes 3 à 13 + parties de 5, 14)"),
        ("Partiel", "SSO/LDAP, intégration compta, consolidation groupe, GED juridique, reporting tutelle normalisé"),
        ("Hors périmètre", "Gouvernance SI (axe 1), Infrastructure (axe 2), Signature électronique (14.8)"),
    ]

    row = 3
    for left, right in sections:
        if right is None and left:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            c = ws.cell(row=row, column=1, value=left)
            c.font = AXIS_FONT
            c.fill = AXIS_FILL
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 22
        elif left or right:
            cl = ws.cell(row=row, column=1, value=left)
            cl.font = Font(bold=True)
            cl.alignment = WRAP_TOP
            cl.border = BORDER

            cr = ws.cell(row=row, column=2, value=right)
            cr.alignment = WRAP_TOP
            cr.border = BORDER

            if left == "Natif":
                cl.fill = SIGA_NATIF_FILL
            elif left == "Partiel":
                cl.fill = SIGA_PARTIEL_FILL
            elif left == "Hors périmètre":
                cl.fill = SIGA_HORS_FILL
        row += 1


def main():
    wb = Workbook()

    ws_q = wb.active
    build_questionnaire_sheet(ws_q)

    ws_s = wb.create_sheet()
    build_synthese_sheet(ws_s)

    ws_e = wb.create_sheet()
    build_etablissement_sheet(ws_e)

    ws_l = wb.create_sheet()
    build_legende_sheet(ws_l)

    # Réordonner pour que Légende ne soit pas en premier
    wb._sheets = [ws_q, ws_s, ws_e, ws_l]

    out = Path(__file__).resolve().parent / "audit_questionnaire_si.xlsx"
    wb.save(out)
    print(f"Fichier généré : {out}")


if __name__ == "__main__":
    main()
