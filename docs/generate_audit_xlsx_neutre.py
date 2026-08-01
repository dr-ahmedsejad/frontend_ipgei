"""
Génère le questionnaire d'audit SI au format Excel (.xlsx) — VERSION NEUTRE.

Aucune référence à SIGA ni à aucun outil commercial.
À remettre directement aux établissements audités.

Sortie : docs/audit_questionnaire_si_neutre.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# Contenu — neutre
# ---------------------------------------------------------------------------

AXES = [
    {
        "num": 1,
        "titre": "Gouvernance & stratégie SI",
        "questions": [
            ("1.1", "Schéma directeur SI formalisé ? Date de dernière révision ?"),
            ("1.2", "Responsable SI désigné (DSI interne / prestataire / volontaire) ?"),
            ("1.3", "Budget annuel SI et part dédiée aux applicatifs métier ?"),
            ("1.4", "Cartographie applicative à jour ?"),
            ("1.5", "Comité de pilotage SI (fréquence, composition) ?"),
        ],
    },
    {
        "num": 2,
        "titre": "Infrastructure, sécurité & conformité",
        "questions": [
            ("2.1", "Hébergement (on-premise / cloud / mixte) et prestataire ?"),
            ("2.2", "Politique de sauvegarde (fréquence, externalisation, tests de restauration documentés) ?"),
            ("2.3", "Authentification : SSO, AD/LDAP, comptes locaux ?"),
            ("2.4", "Gestion des rôles et habilitations formalisée (matrice rôles × modules) ?"),
            ("2.5", "Mécanisme de déblocage de compte tracé ?"),
            ("2.6", "Conformité RGPD / loi locale données personnelles ?"),
            ("2.7", "Politique de mot de passe (rotation, complexité, MFA) ?"),
        ],
    },
    {
        "num": 3,
        "titre": "Audit & traçabilité",
        "questions": [
            ("3.1", "Journal d'audit centralisé append-only (non modifiable) ?"),
            ("3.2", "Timeline par entité (qui a modifié quoi, quand) ?"),
            ("3.3", "Statistiques globales d'activité par action / par objet métier ?"),
            ("3.4", "Export des logs d'audit (CSV, JSON, autre) ?"),
            ("3.5", "Restriction d'accès aux logs (utilisateur ne voit pas l'audit d'un autre) ?"),
            ("3.6", "Historique des modifications visible côté utilisateur final ?"),
        ],
    },
    {
        "num": 4,
        "titre": "Référentiels académiques",
        "questions": [
            ("4.1", "Référentiel filières, départements, niveaux, modules ?"),
            ("4.2", "Versionnage par année universitaire ?"),
            ("4.3", "Gestion semestres, semaines, créneaux horaires ?"),
            ("4.4", "Référentiel jours de la semaine paramétrable ?"),
            ("4.5", "Référentiel salles avec capacité et équipement ?"),
            ("4.6", "Calendrier adapté (Ramadan, événements locaux) ?"),
            ("4.7", "Périodes de réclamation paramétrables ?"),
        ],
    },
    {
        "num": 5,
        "titre": "Multi-établissement",
        "questions": [
            ("5.1", "Plusieurs institutions partagent-elles un même SI ?"),
            ("5.2", "Isolation stricte des données par institution (pas de fuite cross-institution) ?"),
            ("5.3", "Consolidation cross-institution pour direction du groupe ?"),
            ("5.4", "Identifiant unique étudiant / enseignant au niveau groupe ?"),
        ],
    },
    {
        "num": 6,
        "titre": "Pré-inscriptions & admissions",
        "questions": [
            ("6.1", "Pré-inscription en ligne (lien public, formulaire web) ?"),
            ("6.2", "Token sécurisé / lien unique par candidat ?"),
            ("6.3", "Workflow validation candidature → inscription administrative ?"),
            ("6.4", "Dérogations d'inscription tracées ?"),
            ("6.5", "Import en masse étudiants (CSV / Excel) ?"),
        ],
    },
    {
        "num": 7,
        "titre": "Vie étudiante & scolarité",
        "questions": [
            ("7.1", "Dossier étudiant centralisé (état civil, parcours, documents) ?"),
            ("7.2", "Recherche multi-critères étudiants ?"),
            ("7.3", "Inscription pédagogique (choix modules) séparée de l'administrative ?"),
            ("7.4", "Suivi de la progression académique (semestre / année) ?"),
            ("7.5", "Gestion des comptes étudiants (création, blocage, déblocage) ?"),
            ("7.6", "Portail étudiant (notes, EDT, absences, documents) ?"),
        ],
    },
    {
        "num": 8,
        "titre": "Emplois du temps",
        "questions": [
            ("8.1", "Conception des emplois du temps centralisée ou décentralisée par département ?"),
            ("8.2", "Détection automatique des conflits (enseignant / salle / classe) ?"),
            ("8.3", "Gestion des séances de rattrapage et changements de dernière minute ?"),
            ("8.4", "Archivage des emplois du temps par semaine / année ?"),
            ("8.5", "Diffusion via portail enseignant + portail étudiant ?"),
        ],
    },
    {
        "num": 9,
        "titre": "Ressources enseignants & vacations",
        "questions": [
            ("9.1", "Annuaire enseignants distinguant permanents / vacataires / invités ?"),
            ("9.2", "Suivi des charges horaires par enseignant ?"),
            ("9.3", "Détail des enseignements par enseignant (modules, heures, classes) ?"),
            ("9.4", "Génération des états de paiement des vacations ?"),
            ("9.5", "Export bancaire des virements ?"),
            ("9.6", "Intégration avec la comptabilité générale ?"),
        ],
    },
    {
        "num": 10,
        "titre": "Suivi pédagogique & présences",
        "questions": [
            ("10.1", "Pointage de séance (enseignant présent, séance effectivement tenue) ?"),
            ("10.2", "Saisie des absences étudiants par séance / par salle ?"),
            ("10.3", "Import en masse des absences (rétro-saisie depuis fichier) ?"),
            ("10.4", "Gestion des justificatifs d'absence avec workflow ?"),
            ("10.5", "Saisie des absences via portail enseignant ?"),
            ("10.6", "Alertes automatiques au-delà d'un seuil d'absences ?"),
        ],
    },
    {
        "num": 11,
        "titre": "Évaluations, délibérations & qualité examens",
        "questions": [
            ("11.1", "Saisie des notes par enseignant via portail ?"),
            ("11.2", "Saisie anonyme des copies (anonymat préservé) ?"),
            ("11.3", "Sessions normale / rattrapage distinguées ?"),
            ("11.4", "Pondération CC / TP / Examen paramétrable ?"),
            ("11.5", "Compensation modulaire et semestrielle automatisée ?"),
            ("11.6", "Délibérations avec procès-verbaux générés automatiquement ?"),
            ("11.7", "Rachats jury tracés dans le SI ?"),
            ("11.8", "Décisions annuelles codifiées (passage / redoublement / exclusion / année blanche) ?"),
            ("11.9", "Articles réglementaires référencés dans les codes de décision ?"),
        ],
    },
    {
        "num": 12,
        "titre": "Stages & insertion professionnelle",
        "questions": [
            ("12.1", "Conventions de stage numérisées ?"),
            ("12.2", "Statuts workflow (brouillon → soumise → en cours → terminée) ?"),
            ("12.3", "Distinction stage / Projet de Fin d'Études (PFE) ?"),
            ("12.4", "Tuteur académique + tuteur entreprise tracés ?"),
            ("12.5", "Évaluation tri-axiale (note entreprise / rapport / soutenance) ?"),
            ("12.6", "Jury de soutenance composé de plusieurs membres ?"),
            ("12.7", "Validation PFE séparée de la note finale ?"),
            ("12.8", "Classement / palmarès stages ?"),
            ("12.9", "Dérogations médicales avec justificatif fichier ?"),
        ],
    },
    {
        "num": 13,
        "titre": "Réclamations encadrées",
        "questions": [
            ("13.1", "Réclamations étudiants tracées par type (absence / note / autre) ?"),
            ("13.2", "Fenêtres temporelles d'ouverture des réclamations paramétrables ?"),
            ("13.3", "Réclamation rattachée à l'objet contesté (séance / note / session) ?"),
            ("13.4", "Pièce jointe justificative ?"),
            ("13.5", "Workflow de traitement (soumise → en cours → acceptée / rejetée) ?"),
            ("13.6", "Visibilité côté étudiant via portail ?"),
            ("13.7", "Visibilité côté enseignant via portail ?"),
        ],
    },
    {
        "num": 14,
        "titre": "Documents, communication & reporting",
        "questions": [
            ("14.1", "Génération automatique de documents (attestations, relevés de notes) ?"),
            ("14.2", "Dossier documentaire numérique par étudiant ?"),
            ("14.3", "Téléchargement des documents via portail étudiant ?"),
            ("14.4", "Notifications internes (intra-application) ?"),
            ("14.5", "Notifications email / SMS sortantes ?"),
            ("14.6", "Tableaux de bord direction (statistiques) ?"),
            ("14.7", "Reporting réglementaire vers tutelle (export normé) ?"),
            ("14.8", "Signature électronique des documents officiels ?"),
            ("14.9", "Gestion électronique des documents avec archivage légal ?"),
        ],
    },
    {
        "num": 15,
        "titre": "Site web institutionnel & présence numérique",
        "questions": [
            ("15.1", "Site web institutionnel à jour ? Date de la dernière refonte ?"),
            ("15.2", "CMS utilisé (WordPress, Drupal, Strapi, Next.js custom…) maintenu et à jour ?"),
            ("15.3", "Site responsive mobile (consultation principalement smartphone par les candidats) ?"),
            ("15.4", "Site multilingue (FR / AR / EN selon contexte) ?"),
            ("15.5", "Accessibilité numérique (RGAA / WCAG 2.1 niveau AA) ?"),
            ("15.6", "Catalogue des formations à jour avec maquettes / débouchés / fiches filières ?"),
            ("15.7", "Calendrier universitaire publié (vacances, examens, rentrées) ?"),
            ("15.8", "Actualités & agenda événementiel mis à jour régulièrement (≥ mensuel) ?"),
            ("15.9", "Formulaire de pré-candidature en ligne intégré au SI interne (pas de re-saisie) ?"),
            ("15.10", "Annuaire enseignants / responsables / services consultable depuis le site ?"),
            ("15.11", "Lien vers portails (étudiant, enseignant) depuis page d'accueil ?"),
            ("15.12", "Single Sign-On (SSO) depuis le site → portail SI (une seule authentification) ?"),
            ("15.13", "Brochures téléchargeables (PDF par filière) à jour ?"),
            ("15.14", "Témoignages diplômés / alumni mis en avant ?"),
            ("15.15", "Partenariats académiques & entreprises affichés ?"),
            ("15.16", "Espace presse / publications scientifiques / rapports d'activité ?"),
            ("15.17", "Présence et activité sur réseaux sociaux (Facebook, LinkedIn, YouTube, Instagram, TikTok) ?"),
            ("15.18", "Newsletter / mailing institutionnel actif (fréquence, taux d'ouverture suivi) ?"),
            ("15.19", "Mentions légales + politique cookies + politique vie privée (RGPD / loi locale) ?"),
            ("15.20", "Statistiques de fréquentation suivies (Google Analytics, Matomo) ?"),
            ("15.21", "Référencement SEO suivi (positionnement sur requêtes clés) ?"),
            ("15.22", "Hébergement (prestataire, SLA, sauvegarde, certificat SSL) documenté ?"),
            ("15.23", "Charte graphique cohérente entre site, portail SI interne et documents officiels ?"),
        ],
    },
]


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1F4E79")  # bleu neutre

SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="555555")

AXIS_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
AXIS_FILL = PatternFill("solid", fgColor="1F4E79")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")

ROW_ALT_FILL = PatternFill("solid", fgColor="F2F6FA")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


# ---------------------------------------------------------------------------
# Construction du classeur
# ---------------------------------------------------------------------------

def build_questionnaire_sheet(ws):
    ws.title = "Questionnaire"

    ws.merge_cells("A1:D1")
    ws["A1"] = "Questionnaire d'audit du Système d'Information"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    ws["A2"] = "État des lieux des outils de gestion — Établissements supérieurs"
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:D3")
    ws["A3"] = (
        "Notation : 0 = Inexistant · 1 = Manuel/papier · 2 = Tableurs · "
        "3 = Logiciel partiel · 4 = Logiciel intégré · 5 = Intégré + portail utilisateur"
    )
    ws["A3"].font = Font(italic=True, size=10, color="555555")
    ws["A3"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    headers = ["#", "Question", "Note (0-5)", "Observations"]
    header_row = 5
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 28

    widths = {"A": 8, "B": 75, "C": 12, "D": 60}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    dv_note = DataValidation(
        type="whole",
        operator="between",
        formula1=0,
        formula2=5,
        showErrorMessage=True,
        errorTitle="Note invalide",
        error="La note doit être un entier entre 0 et 5.",
    )
    ws.add_data_validation(dv_note)

    row = header_row + 1
    for axe in AXES:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value=f"Axe {axe['num']} — {axe['titre']}")
        cell.font = AXIS_FONT
        cell.fill = AXIS_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 24
        row += 1

        for i, (num, question) in enumerate(axe["questions"]):
            alt = (i % 2 == 1)
            values = [num, question, None, ""]
            for col_idx, val in enumerate(values, start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.border = BORDER
                c.alignment = WRAP_TOP if col_idx in (2, 4) else WRAP_CENTER
                if alt:
                    c.fill = ROW_ALT_FILL

            dv_note.add(ws.cell(row=row, column=3))
            row += 1

    ws.freeze_panes = "A6"


def build_synthese_sheet(ws):
    ws.title = "Synthèse par axe"

    ws.merge_cells("A1:D1")
    ws["A1"] = "Synthèse par axe — après dépouillement"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Axe", "Couverture (0-5)", "Maturité (0-5)", "Risque (faible/moyen/élevé)"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[3].height = 30

    widths = {"A": 55, "B": 18, "C": 18, "D": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    dv_note = DataValidation(
        type="whole", operator="between", formula1=0, formula2=5,
        showErrorMessage=True, errorTitle="Note invalide",
        error="La note doit être un entier entre 0 et 5.",
    )
    ws.add_data_validation(dv_note)

    dv_risque = DataValidation(
        type="list", formula1='"faible,moyen,élevé"',
        showErrorMessage=True, errorTitle="Valeur invalide",
        error="Choisir parmi : faible / moyen / élevé",
    )
    ws.add_data_validation(dv_risque)

    row = 4
    for i, axe in enumerate(AXES):
        alt = (i % 2 == 1)
        label = f"Axe {axe['num']} — {axe['titre']}"
        ws.cell(row=row, column=1, value=label).alignment = WRAP_TOP
        for col_idx in range(1, 5):
            c = ws.cell(row=row, column=col_idx)
            c.border = BORDER
            if alt:
                c.fill = ROW_ALT_FILL
        dv_note.add(ws.cell(row=row, column=2))
        dv_note.add(ws.cell(row=row, column=3))
        dv_risque.add(ws.cell(row=row, column=4))
        row += 1

    ws.freeze_panes = "A4"


def build_etablissement_sheet(ws):
    ws.title = "Établissement"

    ws.merge_cells("A1:B1")
    ws["A1"] = "Identification de l'établissement audité"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    fields = [
        "Nom de l'établissement",
        "Type (public / privé / mixte)",
        "Effectif étudiants",
        "Effectif enseignants (permanents + vacataires)",
        "Nombre de filières",
        "Nombre de départements",
        "Nombre de sites / campus",
        "Tutelle",
        "Nom du répondant principal",
        "Fonction",
        "Date de l'audit",
        "Auditeur",
    ]

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 50

    for i, label in enumerate(fields, start=3):
        c_label = ws.cell(row=i, column=1, value=label)
        c_label.font = Font(bold=True)
        c_label.fill = PatternFill("solid", fgColor="F2F6FA")
        c_label.border = BORDER
        c_label.alignment = WRAP_TOP

        c_val = ws.cell(row=i, column=2, value="")
        c_val.border = BORDER
        c_val.alignment = WRAP_TOP


def build_legende_sheet(ws):
    ws.title = "Légende"

    ws.merge_cells("A1:B1")
    ws["A1"] = "Légende & mode d'emploi"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    sections = [
        ("Grille de notation", None),
        ("0", "Inexistant — aucun processus"),
        ("1", "Manuel / papier uniquement"),
        ("2", "Tableurs (Excel) en local, pas de partage structuré"),
        ("3", "Logiciel partiel — un module isolé, pas d'intégration avec les autres"),
        ("4", "Logiciel intégré — référentiels partagés entre modules"),
        ("5", "Intégré + portail utilisateur (étudiant et/ou enseignant)"),
        ("", ""),
        ("Mode d'emploi", None),
        ("1", "Remplir la feuille « Établissement » avec les informations générales."),
        ("2", "Remplir la feuille « Questionnaire » : note 0-5 + observations par question."),
        ("3", "Remplir la feuille « Synthèse par axe » après dépouillement."),
        ("4", "Pour un nouvel établissement, dupliquer le classeur (clic droit sur l'onglet > Déplacer/Copier)."),
        ("", ""),
        ("Couverture par axe", None),
        ("Couverture (0-5)", "Combien de questions de l'axe ont une réponse satisfaisante."),
        ("Maturité (0-5)", "Qualité globale de la solution actuelle (papier ↔ intégré)."),
        ("Risque", "Impact si l'axe reste non-outillé (faible / moyen / élevé)."),
    ]

    row = 3
    for left, right in sections:
        if right is None and left:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            c = ws.cell(row=row, column=1, value=left)
            c.font = AXIS_FONT
            c.fill = AXIS_FILL
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 22
        elif left or right:
            cl = ws.cell(row=row, column=1, value=left)
            cl.font = Font(bold=True)
            cl.alignment = WRAP_TOP
            cl.border = BORDER

            cr = ws.cell(row=row, column=2, value=right)
            cr.alignment = WRAP_TOP
            cr.border = BORDER
        row += 1


def main():
    wb = Workbook()

    ws_q = wb.active
    build_questionnaire_sheet(ws_q)

    ws_s = wb.create_sheet()
    build_synthese_sheet(ws_s)

    ws_e = wb.create_sheet()
    build_etablissement_sheet(ws_e)

    ws_l = wb.create_sheet()
    build_legende_sheet(ws_l)

    wb._sheets = [ws_q, ws_s, ws_e, ws_l]

    out = Path(__file__).resolve().parent / "audit_questionnaire_si_neutre.xlsx"
    wb.save(out)
    print(f"Fichier généré : {out}")


if __name__ == "__main__":
    main()
